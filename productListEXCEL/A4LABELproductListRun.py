'''
--Excel 文件构成i
    --目录文件>>中转文件(打开编辑不保存,编辑格式 行宽高)>> 保存文件
-----
A4LABEL project
        --自动生成打印列表--

    读取产品文件 return 产品列表所有信息

    获取传递回来的信息
    插入列表
'''
import os
import shutil

#目标 复制原有文件行其中包含图片 to 一个新的文件
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PilImage
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font


class ExcelEdit:
    def __init__(self):
        pass

        #self.full_client_document_path = full_client_document
        #self.full_destination_document_path =  os.path.join(full_destination_document, '排产列表.xlsx')

        #print('in productlist file, full_destination_document_path', self.full_destination_document_path)

    #读取文件获取最大行、最大列 传回main文件用于建立循环
    def A4LABEL_max_row_column(self,clien_file_path):
        # 目标客户文件路径
        print('in A4LABEL_get_information full client path >> ', clien_file_path)
        client_wb = load_workbook(clien_file_path)
        client_ws = client_wb['产品列表']

        client_max_row = client_ws.max_row
        client_max_column = client_ws.max_column
        print('产品列表的最大行数 >>> ', client_max_row)
        print('产品列表的最大列数 >>> ', client_max_column)
        column_size = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                       'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

        return client_max_row, client_max_column
    # 在Label4的main 文件建立循环
    #                               传入文件路径         目标行
    def A4LABEL_get_title_information(self,clien_file_path ):
        #目标客户文件路径
        print('in A4LABEL_get_information full client path >> ', clien_file_path)
        client_wb = load_workbook(clien_file_path)
        client_ws = client_wb['产品列表']

        client_max_row = client_ws.max_row
        client_max_column = client_ws.max_column

        column_size = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                       'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        #获取第一行 图片 名称 数据
        client_file_title_data = []
        for i in range(1, client_max_column+1):
            client_file_title_data.append(client_ws.cell(row=1,column=i).value)

        return client_file_title_data


    #获取指定行信息
    def A4LABEL_get_row_information(self,clien_file_path,target_row, target_column ):
        #目标客户文件路径
        print('in A4LABEL_get_information full client path >> ', clien_file_path)
        client_wb = load_workbook(clien_file_path)
        client_ws = client_wb['产品列表']

        client_max_row = client_ws.max_row
        client_max_column = client_ws.max_column

        column_size = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                       'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        #获取第一行 图片 名称 数据
        client_file_data = []

        client_file_data.append(client_ws.cell(row=target_row,column=target_column).value)

        return client_file_data

    def get_image_for_save(self,clien_file_path,save_pic_path):
        try:
            shutil.rmtree(save_pic_path)
            print(f"文件夹 '{save_pic_path}' 已被删除。")
        except OSError as e:
            print(f"删除文件夹时发生错误: {e.strerror}")

        try:
            os.makedirs(save_pic_path, exist_ok=True)
            print(f"文件夹 '{save_pic_path}' 已被创建（或已存在）。")
        except OSError as e:
            print(f"创建文件夹时发生错误: {e.strerror}")

        client_wb = load_workbook(clien_file_path)
        client_ws = client_wb['产品列表']

        client_max_row = client_ws.max_row
        client_max_column = client_ws.max_column

        for image in client_ws._images:
            #print('锁定的目标图片信息', image.anchor._from)
            #row
            file_name = str(image.anchor._from.row + 1)+'.png'
            #print(file_name)
            temp_pic_name_path = os.path.join(save_pic_path,file_name)
            #print(image.ref)
            #image.height =
            img_pil = PilImage.open(image.ref)

            img_pil.save(temp_pic_name_path)

    def A4_LABEL_init_Excel_file(self,file_path):
        # 锁定排产excel文件
        # 打开
        # 获取最大列
        # 缓存文件
        #获取单元格 行高行宽 格位置
        print('init file >>>>', file_path)

        self.removWorkSheet(file_path)

        destination_wb = load_workbook(file_path)
        destination_ws = destination_wb['产品列表']

        destination_ws.merge_cells('A1:B1')
        destination_ws.merge_cells('D1:E1')
        destination_ws.merge_cells('A7:B7')
        destination_ws.merge_cells('D7:E7')

        destination_ws.row_dimensions[1].height = 187
        destination_ws.row_dimensions[2].height = 39.95
        destination_ws.row_dimensions[3].height = 39.95
        destination_ws.row_dimensions[4].height = 39.95
        destination_ws.row_dimensions[5].height = 39.95
        #destination_ws.row_dimensions[6].height = 39.95
        destination_ws.row_dimensions[6].height = 30
        destination_ws.row_dimensions[7].height = 187
        destination_ws.row_dimensions[8].height = 39.95
        destination_ws.row_dimensions[9].height = 39.95
        destination_ws.row_dimensions[10].height = 39.95
        destination_ws.row_dimensions[11].height = 39.95
        #destination_ws.row_dimensions[13].height = 39.95

        destination_ws.column_dimensions['A'].width = 16
        destination_ws.column_dimensions['B'].width = 30.26
        destination_ws.column_dimensions['C'].width = 6
        destination_ws.column_dimensions['D'].width = 16
        destination_ws.column_dimensions['E'].width = 30.26

        # 创建一个边框样式
        thin = Side(border_style="thin", color="000000")

        # 创建一个边框对象，并设置各个边的样式
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for r in range(2,6):
            for c in range(1,6):
                destination_ws.cell(row=r, column=c).border = border

        for r in range(8,12):
            for c in range(1,6):
                destination_ws.cell(row=r, column=c).border = border

        destination_ws.cell(row=1, column=1).border = border
        destination_ws.cell(row=1, column=2).border = border
        destination_ws.cell(row=1, column=4).border = border
        destination_ws.cell(row=1, column=5).border = border
        #destination_ws.cell(row=1, column=4).border = border
        destination_ws.cell(row=8, column=1).border = border
        destination_ws.cell(row=8, column=2).border = border
        destination_ws.cell(row=8, column=4).border = border
        destination_ws.cell(row=8, column=5).border = border

        # 创建一个Alignment对象，并设置水平居中和垂直居中
        alignment = Alignment(horizontal='center', vertical='center')

        destination_ws.cell(row=2, column=1).value = '厂   家'
        destination_ws.cell(row=3, column=1).value = '产品编号'
        destination_ws.cell(row=4, column=1).value = '产品名称'
        destination_ws.cell(row=5, column=1).value = '生产日期'

        # 将Alignment对象应用到A1单元格
        destination_ws.cell(row=2, column=1).alignment = alignment
        destination_ws.cell(row=3, column=1).alignment = alignment
        destination_ws.cell(row=4, column=1).alignment = alignment
        destination_ws.cell(row=5, column=1).alignment = alignment

        destination_ws.cell(row=2, column=4).value = '厂   家'
        destination_ws.cell(row=3, column=4).value = '产品编号'
        destination_ws.cell(row=4, column=4).value = '产品名称'
        destination_ws.cell(row=5, column=4).value = '生产日期'
        # 将Alignment对象应用到A1单元格
        destination_ws.cell(row=2, column=4).alignment = alignment
        destination_ws.cell(row=3, column=4).alignment = alignment
        destination_ws.cell(row=4, column=4).alignment = alignment
        destination_ws.cell(row=5, column=4).alignment = alignment

        destination_ws.cell(row=8, column=1).value = '厂   家'
        destination_ws.cell(row=9, column=1).value = '产品编号'
        destination_ws.cell(row=10, column=1).value = '产品名称'
        destination_ws.cell(row=11, column=1).value = '生产日期'
        # 将Alignment对象应用到A1单元格
        destination_ws.cell(row=8, column=1).alignment = alignment
        destination_ws.cell(row=9, column=1).alignment = alignment
        destination_ws.cell(row=10, column=1).alignment = alignment
        destination_ws.cell(row=11, column=1).alignment = alignment

        destination_ws.cell(row=8, column=4).value = '厂   家'
        destination_ws.cell(row=9, column=4).value = '产品编号'
        destination_ws.cell(row=10, column=4).value = '产品名称'
        destination_ws.cell(row=11, column=4).value = '生产日期'
        # 将Alignment对象应用到A1单元格
        destination_ws.cell(row=8, column=4).alignment = alignment
        destination_ws.cell(row=9, column=4).alignment = alignment
        destination_ws.cell(row=10, column=4).alignment = alignment
        destination_ws.cell(row=11, column=4).alignment = alignment

        destination_wb.save(file_path)

    def A4_LABEL_insert_Excel_file(self, file_path,select_information):

        print('在点击 A4_LABEL_insert_Excel_file')
        print(file_path)
        print(select_information)

        destination_wb = load_workbook(file_path)
        destination_ws = destination_wb['产品列表']
        print(select_information['window1_pic'])

        if select_information['window1_pic'] != '':
            image_1 = Image(select_information['window1_pic'])  # 替换为你的图片路径
            #image_4 = image_4.resize((350, 170), Image.LANCZOS)
            image_1.anchor = destination_ws.cell(row=1,column=1).coordinate
            destination_ws.add_image(image_1)

        if select_information['window2_pic'] != '':
            image_2 = Image(select_information['window2_pic'])  # 替换为你的图片路径
            #image_4 = image_4.resize((350, 170), Image.LANCZOS)
            image_2.anchor = destination_ws.cell(row=1,column=4).coordinate
            destination_ws.add_image(image_2)


        if select_information['window3_pic'] != '':
            image_3 = Image(select_information['window3_pic'])  # 替换为你的图片路径
            #image_4 = image_4.resize((350, 170), Image.LANCZOS)
            image_3.anchor = destination_ws.cell(row=7,column=1).coordinate
            destination_ws.add_image(image_3)

        if select_information['window4_pic'] != '':
            image_4 = Image(select_information['window4_pic'])  # 替换为你的图片路径
            #image_4 = image_4.resize((350, 170), Image.LANCZOS)
            image_4.anchor = destination_ws.cell(row=7,column=4).coordinate
            destination_ws.add_image(image_4)

        '''self.select_information = {
                    'window1_pic':'','client1':'','productNumber1':'','productName1':'','time1':'',
                    'window2_pic':'', 'client2': '', 'productNumber2': '', 'productName2': '', 'time2': '',
                    'window3_pic': '', 'client3': '', 'productNumber3': '', 'productName3': '', 'time3': '',
                    'window4_pic': '', 'client4': '', 'productNumber4': '', 'productName4': '', 'time4': '',
                }'''
        destination_ws.cell(row=2,column=2).value = select_information['client1']
        destination_ws.cell(row=3, column=2).value = select_information['productNumber1']
        destination_ws.cell(row=4, column=2).value = select_information['productName1']
        destination_ws.cell(row=5,column=2).value = select_information['time1']

        destination_ws.cell(row=2, column=5).value = select_information['client2']
        destination_ws.cell(row=3, column=5).value = select_information['productNumber2']
        destination_ws.cell(row=4, column=5).value = select_information['productName2']
        destination_ws.cell(row=5, column=5).value = select_information['time2']

        destination_ws.cell(row=8, column=2).value = select_information['client3']
        destination_ws.cell(row=9, column=2).value = select_information['productNumber3']
        destination_ws.cell(row=10, column=2).value = select_information['productName3']
        destination_ws.cell(row=11, column=2).value = select_information['time3']

        destination_ws.cell(row=8, column=5).value = select_information['client4']
        destination_ws.cell(row=9, column=5).value = select_information['productNumber4']
        destination_ws.cell(row=10, column=5).value = select_information['productName4']
        destination_ws.cell(row=11, column=5).value = select_information['time4']

        # 创建一个Alignment对象，并设置水平居中和垂直居中
        alignment = Alignment(horizontal='center', vertical='center')

        # 将Alignment对象应用到A1单元格
        destination_ws.cell(row=2, column=2).alignment = alignment
        destination_ws.cell(row=3, column=2).alignment = alignment
        destination_ws.cell(row=4, column=2).alignment = alignment
        destination_ws.cell(row=5, column=2).alignment = alignment

        destination_ws.cell(row=2, column=5).alignment = alignment
        destination_ws.cell(row=3, column=5).alignment = alignment
        destination_ws.cell(row=4, column=5).alignment = alignment
        destination_ws.cell(row=5, column=5).alignment = alignment

        destination_ws.cell(row=8, column=2).alignment = alignment
        destination_ws.cell(row=9, column=2).alignment = alignment
        destination_ws.cell(row=10, column=2).alignment = alignment
        destination_ws.cell(row=11, column=2).alignment = alignment

        destination_ws.cell(row=8, column=5).alignment = alignment
        destination_ws.cell(row=9, column=5).alignment = alignment
        destination_ws.cell(row=10, column=5).alignment = alignment
        destination_ws.cell(row=11, column=5).alignment = alignment



        destination_wb.save(file_path)




    def removWorkSheet(self,file_path):

        # 检查文件是否存在
        if os.path.exists(file_path):
            # 删除文件
            os.remove(file_path)
            print(f"{file_path} 已被删除")
        else:
            print(f"{file_path} 文件不存在")

        # 打开Excel文件
        wb = Workbook(file_path)
        try:
            # 获取工作表对象
            ws = wb['产品列表']
            # 删除工作表
            wb.remove(ws)
            ws = wb.create_sheet('产品列表')
        except:
            ws = wb.create_sheet('产品列表')
        #sheet1  2  3
        try:
            # 获取工作表对象
            ws = wb['Sheet1']
            # 删除工作表
            print('删除工作表1')
            wb.remove(ws)
        except:
            print('没删除工作表1')

        try:
            # 获取工作表对象
            ws = wb['Sheet2']
            wb.remove(ws)
            print('删除工作表2')
        except:
            print('没删除工作表2')

        try:
            # 获取工作表对象
            ws = wb['Sheet3']
            wb.remove(ws)
            print('删除工作表3')
        except:
            print('没删除工作表3')

        # 保存修改后的Excel文件
        wb.save(file_path)

#########################

    def insert_quantity_product(self, list_product_quantity):
        #锁定排产excel文件
        #打开
        #获取最大列
        # 缓存文件

        destination_wb = load_workbook(self.full_destination_document_path)
        destination_ws = destination_wb['产品列表']

        # 获取文件最大行 列
        # 矩阵循环
        destination_max_row = destination_ws.max_row
        destination_max_column = destination_ws.max_column
        print('产品列表的最大行数 >>> ', destination_max_row)
        print('产品列表的最大列数 >>> ', destination_max_column)
        column_size = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                       'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        try:
            for i in range(2,len(list_product_quantity)+2):
                destination_ws.cell(row=i, column=destination_max_column+1).value = list_product_quantity[i-2]
                destination_ws[column_size[destination_max_column+1] + str(i)].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                # 保存更改
            destination_wb.save(self.full_destination_document_path)
        except Exception as e:
            print('产品数量并未写入')
            print(f"捕捉到异常: {e}")
            destination_wb.save(self.full_destination_document_path)



    #输入元素(客户产品列表、产品编号)
    def strtRunFile(self,customer_product_list,product_number):
        #判断文件里是否有该数据
        have_data = False
        # 加载现有的Excel文件
        client_document = os.path.join(self.full_client_document_path, customer_product_list)
        source_wb = load_workbook(client_document) #'宇森.xlsx'
        source_worksheet = source_wb['产品列表']

        #缓存文件

        destination_wb =load_workbook(self.full_destination_document_path)
        destination_ws = destination_wb['产品目录']

                #目标 编号
        search_value = product_number #'4220106'
        destination_max_roww = destination_ws.max_row + 1 #缓存文件最大行
        destination_max_column = destination_ws.max_column
                   # 查找数据
                   #获取文件最大行 列
                   #矩阵循环
        source_max_row = source_worksheet.max_row
        source_max_column = source_worksheet.max_column
        print('目标 客户产品列表的最大行数 >>> ',source_max_row)
        print('目标 客户产品列表的最大列数 >>> ',source_max_column)
        column_size = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']



                    #两个迭代循环 横向 与 竖向，形成矩阵迭代
        for row in range(1,source_max_row+2):
            for column in range(1,source_max_column):
                        #矩阵中横竖定位，形成cell的定位，根据内容判别
                if source_worksheet.cell(row,column).value == search_value:
                    print('匹配到了搜索目标 >>> ', search_value)


                                #目标行的所有内容历遍
                    for pick_column in range(1,source_max_column):
                            # print(excel_sheet.cell(row = row, column=pick_column).value)
                        row_height = source_worksheet.row_dimensions[row].height #复制行高
                        col_width = source_worksheet.column_dimensions[column_size[pick_column]].width  #复制列宽
                        target_row = row  #矩阵中历遍的 锁中的行
                        target_column_in_letter = column_size[pick_column]#目标列 与字母进行互换
                        target_row_height = row_height
                        target_column_width = col_width
                        target_column = pick_column
                        target_cell = target_column_in_letter + str(destination_max_roww)
                                    # 设置第一列的宽度为20
                                    #复制行高 列宽
                                    #指定单元格
                        destination_ws.column_dimensions[target_column_in_letter].width = col_width
                        destination_ws.row_dimensions[destination_max_roww].height = row_height

                        destination_ws.cell(row=destination_max_roww, column=pick_column).value = source_worksheet.cell(row,pick_column).value
                        #align = Alignment(horizontal='center', vertical='center')
                        #destination_ws[column_size[pick_column] + str(destination_max_roww)].alignment = align
                        destination_ws[column_size[pick_column] + str(destination_max_roww)].alignment = Alignment(
                                wrapText=True, horizontal='center', vertical='center')

        '''           
        for row in range(1, destination_max_roww + 2):
            for column in range(1, destination_max_column):
                if source_worksheet.cell(row, column).value == search_value:
                    print('匹配到了搜索目标 >>> ', search_value)
                    print("没有匹配到的产品",search_value)
                    have_data = True
                    continue
        if have_data:
            pass
        else:
            print(search_value)
            print(destination_ws.cell(row=destination_max_roww, column=1).value)
            '''
        destination_ws.cell(row=destination_max_roww, column=1).value = search_value
        try:
            for image in source_worksheet._images:
                if image.anchor._from.row == int(target_row) - 1:
                    print('锁定的目标图片信息', image.anchor._from)
                        #temp = image.anchor
                    pic_column = column_size[image.anchor._from.col + 1]
                        # 锁定目标图片 插入目标文件
                    image.anchor = destination_ws.cell(row= destination_max_roww , column= image.anchor._from.col + 1 ).coordinate
                    print('图片的位置',)
                        # 设置图片尺寸 宽度小于格子 高度为宽的0.8
                        # image.width = 100
                    image.width = source_worksheet.column_dimensions[pic_column].width * 7
                        # image.height = 80
                    #image.height = image.width * 0.8
                    image.height = image.height * 0.8
                    destination_ws.add_image(image)

                        #image.anchor = temp



                                # 复制行 目标单独一行目标行 复制
                        #source_row_cells = list(source_worksheet[row])  # 需要复制的目标行
                        #for cell in source_row_cells:
                         #   destination_ws.cell(row = destination_max_roww, column=cell.column).value = cell.value
                                # 创建一个居中对齐的Alignment对象
                            #align = Alignment(horizontal='center', vertical='center')
                            #destination_ws[column_size[cell.column]+str(destination_max_roww)].alignment = align


            #print('行高',row_height)
            #print('列宽', col_width)
            #print('客户目标 格', target_column_in_letter+ str(target_row))

                #锁定目标图片
                #根据图片所在行判定位置，记得加1
            #for picRow in pic_row:
                #for image in source_worksheet._images:
                    #print(image.anchor._from)
                        # print(image.anchor._from.col)
                        # print(image.anchor._from.row)

                    #if image.anchor._from.row == int(picRow)-1:
                        #print('锁定的目标图片信息', image.anchor._from)
                        #temp = image.anchor
                        #pic_column = column_size[image.anchor._from.col + 1]
                            #锁定目标图片 插入目标文件
                        #image.anchor = destination_ws.cell(row= destination_max_roww , column= image.anchor._from.col + 1 ).coordinate
                            #设置图片尺寸 宽度小于格子 高度为宽的0.8
                            #image.width = 100
                        #image.width = source_worksheet.column_dimensions[pic_column].width * 7
                            #image.height = 80
                        #image.height = image.width *0.8
                        #destination_ws.add_image(image)

                        #image.anchor = temp


            #    destination_ws.append(row)
            # 保存更改
            destination_wb.save(self.full_destination_document_path)
        except:
            destination_wb.save(self.full_destination_document_path)


    def temp(self):
        source_wb = load_workbook('启新-南华密封圈模具报价2024517.xlsx')  #
        source_worksheet = source_wb['产品列表']

        source_max_row = source_worksheet.max_row

        tempValue = []
        print('文件的最大行数', source_max_row)
        for i in range(2,source_max_row+1):
            tempValue.append(source_worksheet.cell(i, 9).value)

        return tempValue


if __name__ == "__main__":

    temppath = 'C:\\Users\\hello\\Desktop\\A4LabelFor4\\document\\creatnewExcelfile\\A4LabelExcel.xlsx'
    temppicpath = 'C:\\Users\\hello\\Desktop\\A4LabelFor4\\document\\pinPicture\\temppic'
    temp = ExcelEdit()
    temp.removWorkSheet(temppath)
    temp.A4_LABEL_init_Excel_file(temppath)
    #temp.A4_LABEL_insert_Excel_file('C:\\Users\\hello\\Desktop\\A4LabelFor4\\document\\pinPicture',)





