'''
--Excel 文件构成i
    --目录文件>>中转文件(打开编辑不保存,编辑格式 行宽高)>> 保存文件
-----
封装
    能多次大批量处理
    输入
        产品编号 以list[]形式输入
    输出
        缓存文件
            每次操作都会被重新覆盖
'''
import os

#目标 复制原有文件行其中包含图片 to 一个新的文件
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment


class ExcelEdit:
    def __init__(self,full_client_document, full_destination_document):
        self.full_client_document_path = full_client_document
        self.full_destination_document_path =  os.path.join(full_destination_document, '排产列表.xlsx')

        print('in productlist file, full_destination_document_path', self.full_destination_document_path)

    def insert_quantity_product(self, list_product_quantity):
        #锁定排产excel文件
        #打开
        #获取最大列
        # 缓存文件

        destination_wb = load_workbook(self.full_destination_document_path)
        destination_ws = destination_wb['产品目录']

        # 获取文件最大行 列
        # 矩阵循环
        destination_max_row = destination_ws.max_row
        destination_max_column = destination_ws.max_column
        print('产品列表的最大行数 >>> ', destination_max_row)
        print('产品列表的最大列数 >>> ', destination_max_column)
        column_size = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                       'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        try:
            for i in range(2,len(list_product_quantity)+2):
                destination_ws.cell(row=i, column=destination_max_column+1).value = list_product_quantity[i-2]
                destination_ws[column_size[destination_max_column+1] + str(i)].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                # 保存更改
            destination_wb.save(self.full_destination_document_path)
        except Exception as e:
            print('产品数量并未写入')
            print(f"捕捉到异常: {e}")
            destination_wb.save(self.full_destination_document_path)


    #输入元素(客户产品列表、产品编号)
    def strtRunFile(self,customer_product_list,product_number):
        #判断文件里是否有该数据
        have_data = False
        # 加载现有的Excel文件
        client_document = os.path.join(self.full_client_document_path, customer_product_list)
        source_wb = load_workbook(client_document) #'宇森.xlsx'
        source_worksheet = source_wb['产品列表']

        #缓存文件

        destination_wb =load_workbook(self.full_destination_document_path)
        destination_ws = destination_wb['产品目录']

                #目标 编号
        search_value = product_number #'4220106'
        destination_max_roww = destination_ws.max_row + 1 #缓存文件最大行
        destination_max_column = destination_ws.max_column
                   # 查找数据
                   #获取文件最大行 列
                   #矩阵循环
        source_max_row = source_worksheet.max_row
        source_max_column = source_worksheet.max_column
        print('目标 客户产品列表的最大行数 >>> ',source_max_row)
        print('目标 客户产品列表的最大列数 >>> ',source_max_column)
        column_size = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']



                    #两个迭代循环 横向 与 竖向，形成矩阵迭代
        for row in range(1,source_max_row+2):
            for column in range(1,source_max_column):
                        #矩阵中横竖定位，形成cell的定位，根据内容判别
                if source_worksheet.cell(row,column).value == search_value:
                    print('匹配到了搜索目标 >>> ', search_value)


                                #目标行的所有内容历遍
                    for pick_column in range(1,source_max_column):
                            # print(excel_sheet.cell(row = row, column=pick_column).value)
                        row_height = source_worksheet.row_dimensions[row].height #复制行高
                        col_width = source_worksheet.column_dimensions[column_size[pick_column]].width  #复制列宽
                        target_row = row  #矩阵中历遍的 锁中的行
                        target_column_in_letter = column_size[pick_column]#目标列 与字母进行互换
                        target_row_height = row_height
                        target_column_width = col_width
                        target_column = pick_column
                        target_cell = target_column_in_letter + str(destination_max_roww)
                                    # 设置第一列的宽度为20
                                    #复制行高 列宽
                                    #指定单元格
                        destination_ws.column_dimensions[target_column_in_letter].width = col_width
                        destination_ws.row_dimensions[destination_max_roww].height = row_height

                        destination_ws.cell(row=destination_max_roww, column=pick_column).value = source_worksheet.cell(row,pick_column).value
                        #align = Alignment(horizontal='center', vertical='center')
                        #destination_ws[column_size[pick_column] + str(destination_max_roww)].alignment = align
                        destination_ws[column_size[pick_column] + str(destination_max_roww)].alignment = Alignment(
                                wrapText=True, horizontal='center', vertical='center')

        '''           
        for row in range(1, destination_max_roww + 2):
            for column in range(1, destination_max_column):
                if source_worksheet.cell(row, column).value == search_value:
                    print('匹配到了搜索目标 >>> ', search_value)
                    print("没有匹配到的产品",search_value)
                    have_data = True
                    continue
        if have_data:
            pass
        else:
            print(search_value)
            print(destination_ws.cell(row=destination_max_roww, column=1).value)
            '''
        destination_ws.cell(row=destination_max_roww, column=1).value = search_value
        try:
            for image in source_worksheet._images:
                if image.anchor._from.row == int(target_row) - 1:
                    print('锁定的目标图片信息', image.anchor._from)
                        #temp = image.anchor
                    pic_column = column_size[image.anchor._from.col + 1]
                        # 锁定目标图片 插入目标文件
                    image.anchor = destination_ws.cell(row= destination_max_roww , column= image.anchor._from.col + 1 ).coordinate
                    print('图片的位置',)
                        # 设置图片尺寸 宽度小于格子 高度为宽的0.8
                        # image.width = 100
                    image.width = source_worksheet.column_dimensions[pic_column].width * 7
                        # image.height = 80
                    image.height = image.width * 0.8
                    destination_ws.add_image(image)

                        #image.anchor = temp



                                # 复制行 目标单独一行目标行 复制
                        #source_row_cells = list(source_worksheet[row])  # 需要复制的目标行
                        #for cell in source_row_cells:
                         #   destination_ws.cell(row = destination_max_roww, column=cell.column).value = cell.value
                                # 创建一个居中对齐的Alignment对象
                            #align = Alignment(horizontal='center', vertical='center')
                            #destination_ws[column_size[cell.column]+str(destination_max_roww)].alignment = align


            #print('行高',row_height)
            #print('列宽', col_width)
            #print('客户目标 格', target_column_in_letter+ str(target_row))

                #锁定目标图片
                #根据图片所在行判定位置，记得加1
            #for picRow in pic_row:
                #for image in source_worksheet._images:
                    #print(image.anchor._from)
                        # print(image.anchor._from.col)
                        # print(image.anchor._from.row)

                    #if image.anchor._from.row == int(picRow)-1:
                        #print('锁定的目标图片信息', image.anchor._from)
                        #temp = image.anchor
                        #pic_column = column_size[image.anchor._from.col + 1]
                            #锁定目标图片 插入目标文件
                        #image.anchor = destination_ws.cell(row= destination_max_roww , column= image.anchor._from.col + 1 ).coordinate
                            #设置图片尺寸 宽度小于格子 高度为宽的0.8
                            #image.width = 100
                        #image.width = source_worksheet.column_dimensions[pic_column].width * 7
                            #image.height = 80
                        #image.height = image.width *0.8
                        #destination_ws.add_image(image)

                        #image.anchor = temp


            #    destination_ws.append(row)
            # 保存更改
            destination_wb.save(self.full_destination_document_path)
        except:
            destination_wb.save(self.full_destination_document_path)

    def removWorkSheet(self):
        # 打开Excel文件
        wb = load_workbook(self.full_destination_document_path)

        try:
            # 获取工作表对象
            ws = wb['产品目录']


            # 删除工作表
            wb.remove(ws)

            ws = wb.create_sheet('产品目录')
        except:
            ws = wb.create_sheet('产品目录')


        # 保存修改后的Excel文件
        wb.save(self.full_destination_document_path)
    def temp(self):
        source_wb = load_workbook('启新-南华密封圈模具报价2024517.xlsx')  #
        source_worksheet = source_wb['产品列表']

        source_max_row = source_worksheet.max_row

        tempValue = []
        print('文件的最大行数', source_max_row)
        for i in range(2,source_max_row+1):
            tempValue.append(source_worksheet.cell(i, 9).value)

        return tempValue


if __name__ == "__main__":
    temp = ExcelEdit()
    target_list = []

#统一 读取的产品列表 排列格式、表格的名称
#插入功能 修改读取的客户文件
#获取信息 修改读取编号的 列数，开始有用信息的那行

    while True:
        CMD = input('要执行的命令: ')
        if CMD == 'd':
            temp.removWorkSheet()
        elif CMD == 'c':
                #target_list = ['4250747','4351195','4260257','4220266','4260257','4250946','4250855','4420291']
        #缓存文件清空
            for i in target_list:
                print(i)
                temp.strtRunFile('启新-南华密封圈模具报价2024517.xlsx',i)#['4250747','4351195','4260257','4250946','4250855','4250945']

        elif CMD == 'g':
            target_list = temp.temp()
            print(target_list)




