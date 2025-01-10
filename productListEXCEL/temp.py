from openpyxl import load_workbook
from openpyxl import workbook

 # 加载Excel文件
excel_file = load_workbook('宇森.xlsx')
excel_sheet = excel_file['产品列表']

search_value = '4220106'
   
   # 查找数据
#获取文件最大行 列
#矩阵循环
max_row = excel_sheet.max_row
max_column = excel_sheet.max_column
print(max_row)
print(max_column)
column_size = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

for row in range(1,max_row):
    for column in range(1,max_column):
        #print(excel_sheet.cell(row,column).value)
        if excel_sheet.cell(row,column).value == search_value:
            #print(excel_sheet.cell(row,column).value)
            for pick_column in range(1,max_column):
                print(excel_sheet.cell(row = row, column=pick_column).value) 
                row_height = excel_sheet.row_dimensions[row].height
                col_width = excel_sheet.column_dimensions[column_size[pick_column]].width

'''    
   print(f"第一列的宽度是: {col_width}")



   if search_value in df['your_column_name'].values:
           print(f'找到数据: {search_value}')
       else:
               print(f'未找到数据: {search_value}')

m openpyxl import load_workbook
 
 # 加载现有的Excel文件
 workbook = load_workbook('example.xlsx')
  
  # 获取工作表
  sheet = workbook.active  # 或者 workbook['工作表名称']
   
   # 获取最大行数
   max_row = sheet.max_row
    
    print(f"最大行数是: {max_row}")
'''
